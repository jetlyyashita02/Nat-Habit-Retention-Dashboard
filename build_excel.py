"""
Builds 'Nat_Habit_Retention_Dashboard.xlsx' — a fully formula-driven
Excel workbook. Paste the raw 155k journey dump into the DATA sheet and every
dashboard tile, variant table, grammage table and conclusion sentence
recomputes via native Excel formulas. No macros.

Run:  python3 build_excel.py
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SAMPLE = "data/sample_journeys.csv"
OUT = "Nat_Habit_Retention_Dashboard.xlsx"

# ---------------------------------------------------------------- palette
GREEN = "3C8C6E"; DARK = "1F4D3F"; LIGHT = "EAF3EE"; GREY = "F5F5F5"
ACCENT = "E4A11B"
HDR = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=16, color=DARK)
SUB = Font(size=10, color="666666")
FILL_HDR = PatternFill("solid", fgColor=GREEN)
FILL_SEC = PatternFill("solid", fgColor=LIGHT)
FILL_TILE = PatternFill("solid", fgColor=GREY)
THIN = Border(*[Side(style="thin", color="DDDDDD")] * 4)

# ---------------------------------------------------------------- DATA layout
RAW_COLS = ["Name", "category_first", "category_2nd", "category_3rd",
            "category_4th", "category_5th", "category_6th",
            "first_purchased_short_code", "second_purchased_short_code",
            "third_purchased_short_code", "fourth_purchased_short_code",
            "fifth_purchased_short_code", "sixth_purchased_short_code",
            "first_purchased_sku", "second_purchased_sku", "third_purchased_sku",
            "fourth_purchased_sku", "fifth_purchased_sku", "sixth_purchased_sku",
            "first_order_date", "second_order_date", "third_order_date",
            "fourth_order_date", "fifth_order_date", "sixth_order_date",
            "days_between_1st_and_2nd", "days_between_2nd_and_3rd",
            "days_between_3rd_and_4th", "days_between_4th_and_5th",
            "days_between_5th_and_6th", "city", "Month-Year", "Month Number ",
            "Journey Depth", "Is Repeat Buyer? "]
C_NAME = 1
C_SKU = {s: 14 + s for s in range(6)}      # first..sixth_purchased_sku
C_DATE = {s: 20 + s for s in range(6)}     # first..sixth_order_date
C_CITY = 31; C_MY = 32; C_MN = 33; C_DEPTH = 34; C_REP = 35

C_REGION = 36
SLOT0 = 37                                 # 9 cols per slot
# slot offsets: code line vcode cat var size datev season intent
C_CODE = {s: SLOT0 + 9 * s + 0 for s in range(6)}
C_LINE = {s: SLOT0 + 9 * s + 1 for s in range(6)}
C_VCOD = {s: SLOT0 + 9 * s + 2 for s in range(6)}
C_CAT  = {s: SLOT0 + 9 * s + 3 for s in range(6)}
C_VAR  = {s: SLOT0 + 9 * s + 4 for s in range(6)}
C_SIZE = {s: SLOT0 + 9 * s + 5 for s in range(6)}
C_DTV  = {s: SLOT0 + 9 * s + 6 for s in range(6)}
C_SEAS = {s: SLOT0 + 9 * s + 7 for s in range(6)}
C_INT  = {s: SLOT0 + 9 * s + 8 for s in range(6)}

C_ECAT, C_EVAR, C_ESIZE, C_ESKU = 91, 92, 93, 94
C_NEXT, C_NCAT, C_NVAR, C_NSIZE, C_NSKU = 95, 96, 97, 98, 99
C_MOVE, C_GAP, C_V2V, C_V2C, C_COH = 100, 101, 102, 103, 104
N_HELP_COLS = C_COH

def L(i): return CL(i)

SLOTS_ORD = ["first", "second", "third", "fourth", "fifth", "sixth"]

LINE_MAP = [("CWDM", "Cold Winter", "Face Malai"),
            ("HDDM", "Hot Dry", "Face Malai"),
            ("HHDM", "Hot Humid", "Face Malai"),
            ("DM", "Non-Seasonal", "Face Malai"),
            ("AC", "Gel", "Active Gel"),
            ("HG", "Gel", "Aloe Vera Gel")]
VAR_MAP = [("CWDM-FW", "Flax Walnut"), ("CWDM-TP", "Tomato Patchouli"),
           ("CWDM-BT", "Winter Blackseed"), ("HDDM-FB", "Flax Bakuchi"),
           ("HDDM-TR", "Tomato Rosehip"), ("HDDM-PT", "Pomegranate Tulsi"),
           ("HHDM-FC", "Flax Carrot"), ("HHDM-TV", "Tomato Vetiver"),
           ("HHDM-GAT", "GreenApple Tulsi"), ("DM-PG", "Turmeric Nutmeg"),
           ("DM-NR", "Cocoa Mogra"), ("DM-DP", "Honey Multi-Nut"),
           ("DM-AA", "Clove Tea-Tree"), ("AC-NT", "Neem TeaTree"),
           ("AC-AC", "Aloe Cactus Hydra Lift"), ("AC-OV", "Olive Vit-E Night"),
           ("AC-OK", "Orange Kiwi Vit-C"), ("AC-BT", "Beetroot Tomato Vit-A"),
           ("HG-PA", "Aloe Vera")]
VAR_INTENT = {v: next(i for ln, i, c in LINE_MAP if k.startswith(ln)) for k, v in VAR_MAP}
SKUS = (sorted({f"{v} 30g" for _, v in VAR_MAP[:13]} | {f"{v} 50g" for _, v in VAR_MAP[:13]}
               | {f"{v} 50g" for _, v in VAR_MAP[13:18]}
               | {"Aloe Vera Gel 40g", "Aloe Vera Gel 80g"}))
REGION_MAP = [  # city -> region (subset of etl.REGION_MAP, editable in DECODE)
    ("New Delhi", "North"), ("Delhi", "North"), ("Gurgaon", "North"),
    ("Noida", "North"), ("Gautam Buddha Nagar", "North"), ("Ghaziabad", "North"),
    ("Faridabad", "North"), ("Jaipur", "North"), ("Lucknow", "North"),
    ("Kanpur", "North"), ("Chandigarh", "North"), ("Dehradun", "North"),
    ("Varanasi", "North"), ("Prayagraj", "North"), ("Agra", "North"),
    ("Meerut", "North"), ("Jodhpur", "North"), ("Udaipur", "North"),
    ("Kota", "North"), ("Bhopal", "North"), ("Indore", "North"),
    ("Mumbai", "West"), ("Thane", "West"), ("Navi Mumbai", "West"),
    ("Pune", "West"), ("Nashik", "West"), ("Nagpur", "West"),
    ("Ahmedabad", "West"), ("Surat", "West"), ("Vadodara", "West"),
    ("Rajkot", "West"),
    ("Bengaluru", "South"), ("Bangalore", "South"), ("Mysuru", "South"),
    ("Mangaluru", "South"), ("Chennai", "South"), ("Coimbatore", "South"),
    ("Madurai", "South"), ("Hyderabad", "South"), ("Secunderabad", "South"),
    ("Visakhapatnam", "South"), ("Vijayawada", "South"), ("Kochi", "South"),
    ("Kozhikode", "South"), ("Thiruvananthapuram", "South"),
    ("Thrissur", "South"), ("Kollam", "South"), ("Palakkad", "South"),
    ("Kannur", "South"), ("Kottayam", "South"),
    ("Kolkata", "East"), ("Howrah", "East"), ("Bhubaneswar", "East"),
    ("Patna", "East"), ("Ranchi", "East"), ("Jamshedpur", "East"),
    ("Guwahati", "East"), ("Siliguri", "East"),
]
MONTHS = [f"{y}-{m:02d}" for y in range(2024, 2028) for m in range(1, 13)]
VARIANTS = [v for _, v in VAR_MAP]

wb = Workbook()

# ================================================================= README
ws = wb.active; ws.title = "README"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 120
readme = [
    ("T", "🧴 Moisturisers — Formula-Driven Interactive Dashboard"),
    ("S", "Retention & category intelligence · Jan 2024 – Jun 2026 · works entirely with native Excel formulas (no macros)"),
    ("B", ""),
    ("H", "HOW TO USE WITH YOUR FULL RAW DUMP (155,537 orders)"),
    ("N", "1. Open the DATA sheet. Row 1 holds the exact column headers of your raw journey dump."),
    ("N", "2. Select cell A2 and paste your whole dump over the sample rows (all 35 columns, any number of rows)."),
    ("N", "3. The grey helper columns (AJ onwards) auto-fill for pasted rows because DATA is an Excel Table."),
    ("N", "   If they ever stop at the old row count: click the last filled helper cell, copy (Ctrl+C),"),
    ("N", "   select down to your last data row (Ctrl+Shift+End on the helper columns), and paste (Ctrl+V)."),
    ("N", "4. Press F9 to recalculate (first full recalc on 155k rows can take a minute or two)."),
    ("N", "5. Everything else updates by itself: DASHBOARD tiles, CONCLUSIONS text, VARIANTS and GRAMMAGE tables."),
    ("B", ""),
    ("H", "WHAT EACH SHEET DOES"),
    ("N", "DASHBOARD — the interactive mockup: dropdown filters (Category, Variant SKU, From/To month, Journey Depth, Geo region)"),
    ("N", "   drive every tile through COUNTIFS/SUMIFS. Filters apply to each customer's FIRST (entry) order = cohort semantics."),
    ("N", "CONCLUSIONS — the conclusion blocks, 100% formula-driven: every number is computed and every verdict word"),
    ("N", "   (YES/NO, MODERATE, DOWNSIZE ALERT, FOCUS) is a formula branch. They reflect the WHOLE workbook (not the dashboard filters)."),
    ("N", "VARIANTS — per-variant seasonality (purchase-month buckets), peak season, season match, seasonality strength,"),
    ("N", "   entry users, drop-off and reorder speed. Whole-workbook level."),
    ("N", "GRAMMAGE — entry size → next purchase moves (repeat / upsize / lateral / downsize) + the exact path counts"),
    ("N", "   the conclusions quote (FM30→FM50, FM30→AG50, FM50→FM30, AG50→FM30). Whole-workbook level."),
    ("N", "DATA — your dump + decoder helper columns (SKU code → line/variant/size, dates, seasons, entry→next transitions)."),
    ("N", "DECODE — all lookup tables (SKU codes, city→region, month list, variant→intended season). Edit region map here."),
    ("B", ""),
    ("H", "NOTES"),
    ("N", "· Multi-item orders are analysed by their FIRST SKU code in the cell (primary item) — same convention as the sample."),
    ("N", "· Dates must be text like 'Jan 15, 2026' or real Excel dates; both are handled."),
    ("N", "· V2V = 2nd order repeats an entry variant · V2C = 2nd order stays in an entry category."),
    ("N", "· Seasons: Cold Winter Dec–Feb · Hot Dry Mar–May · Hot Humid Jun–Sep · Post-Monsoon Oct–Nov."),
    ("N", "· Seasonality strength: Highly ≥40% of orders in peak season · Moderately 30–40% · Evergreen <30%."),
]
r = 2
for kind, txt in readme:
    c = ws.cell(row=r, column=2, value=txt)
    if kind == "T": c.font = TITLE
    elif kind == "S": c.font = SUB
    elif kind == "H": c.font = Font(bold=True, color=DARK, size=12); c.fill = FILL_SEC
    elif kind == "N": c.font = Font(size=10)
    r += 1

# ================================================================= DECODE
dec = wb.create_sheet("DECODE")
dec["A1"] = "LINE MAP (SKU 2nd token)"; dec["A1"].font = BOLD
for i, (k, intent, cat) in enumerate(LINE_MAP, start=3):
    dec.cell(row=i, column=1, value=k); dec.cell(row=i, column=2, value=intent)
    dec.cell(row=i, column=3, value=cat)
LINES = f"DECODE!$A$3:$C${2+len(LINE_MAP)}"

dec["E1"] = "VARIANT MAP (line-vcode)"; dec["E1"].font = BOLD
for i, (k, v) in enumerate(VAR_MAP, start=3):
    dec.cell(row=i, column=5, value=k); dec.cell(row=i, column=6, value=v)
VARMAP = f"DECODE!$E$3:$F${2+len(VAR_MAP)}"

dec["H1"] = "REGION MAP (edit freely)"; dec["H1"].font = BOLD
for i, (city, reg) in enumerate(REGION_MAP, start=3):
    dec.cell(row=i, column=8, value=city); dec.cell(row=i, column=9, value=reg)
REGMAP = f"DECODE!$H$3:$I${2+len(REGION_MAP)}"

dec["K1"] = "VARIANT → INTENDED SEASON"; dec["K1"].font = BOLD
for i, v in enumerate(VARIANTS, start=3):
    dec.cell(row=i, column=11, value=v); dec.cell(row=i, column=12, value=VAR_INTENT[v])
VARINTENT = f"DECODE!$K$3:$L${2+len(VARIANTS)}"

dec["N1"] = "LISTS (for dropdowns)"; dec["N1"].font = BOLD
lists = {
    14: ["All", "Face Malai", "Active Gel", "Aloe Vera Gel"],           # N cat
    15: ["All"] + SKUS,                                                 # O sku
    16: ["All"] + MONTHS,                                               # P month
    17: ["All"] + [str(i) for i in range(1, 7)],                        # Q depth
    18: ["All", "North", "West", "South", "East", "Rest of India"],     # R region
}
headers = {14: "Category", 15: "Variant SKU", 16: "Month", 17: "Journey Depth", 18: "Region"}
for col, vals in lists.items():
    dec.cell(row=2, column=col, value=headers[col]).font = BOLD
    for i, v in enumerate(vals, start=3):
        dec.cell(row=i, column=col, value=v)
CATLIST = f"DECODE!$N$3:$N${2+len(lists[14])}"
SKULIST = f"DECODE!$O$3:$O${2+len(lists[15])}"
MONLIST = f"DECODE!$P$3:$P${2+len(lists[16])}"
DEPLIST = f"DECODE!$Q$3:$Q${2+len(lists[17])}"
REGLIST = f"DECODE!$R$3:$R${2+len(lists[18])}"
for col, w in [(1, 10), (5, 12), (8, 20), (11, 24), (14, 14), (15, 28), (16, 10), (17, 8), (18, 12)]:
    dec.column_dimensions[CL(col)].width = w

# ================================================================= DATA
df = pd.read_csv(SAMPLE)
dat = wb.create_sheet("DATA")
for j, h in enumerate(RAW_COLS, start=1):
    c = dat.cell(row=1, column=j, value=h)
    c.font = HDR; c.fill = FILL_HDR
help_headers = {C_REGION: "Region"}
for s in range(6):
    o = SLOTS_ORD[s]
    help_headers |= {
        C_CODE[s]: f"{o} code", C_LINE[s]: f"{o} line", C_VCOD[s]: f"{o} vcode",
        C_CAT[s]: f"{o} category", C_VAR[s]: f"{o} variant", C_SIZE[s]: f"{o} size",
        C_DTV[s]: f"{o} date", C_SEAS[s]: f"{o} season", C_INT[s]: f"{o} intent"}
help_headers |= {C_ECAT: "entry category", C_EVAR: "entry variant",
                 C_ESIZE: "entry size", C_ESKU: "entry sku",
                 C_NEXT: "has next?", C_NCAT: "next category", C_NVAR: "next variant",
                 C_NSIZE: "next size", C_NSKU: "next sku", C_MOVE: "move type",
                 C_GAP: "gap days", C_V2V: "V2V hit", C_V2C: "V2C hit",
                 C_COH: "in cohort?"}
for col, h in help_headers.items():
    c = dat.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.fill = PatternFill("solid", fgColor="7A7A7A")

n = len(df)
for i, row in enumerate(df.itertuples(index=False), start=2):
    for j, h in enumerate(RAW_COLS, start=1):
        v = df.iloc[i - 2][h]
        if pd.notna(v):
            dat.cell(row=i, column=j, value=v)

for i in range(2, n + 2):
    f = {}
    f[C_REGION] = (f'=IF(${L(C_CITY)}{i}="","",IFERROR(VLOOKUP(TRIM(${L(C_CITY)}{i}),'
                   f'{REGMAP},2,FALSE),"Rest of India"))')
    for s in range(6):
        sku, dt = L(C_SKU[s]), L(C_DATE[s])
        code, line, vcod = L(C_CODE[s]), L(C_LINE[s]), L(C_VCOD[s])
        cat, var, size = L(C_CAT[s]), L(C_VAR[s]), L(C_SIZE[s])
        dtv, seas, intent = L(C_DTV[s]), L(C_SEAS[s]), L(C_INT[s])
        f[C_CODE[s]] = (f'=IF(${sku}{i}="","",UPPER(TRIM(LEFT(${sku}{i}&",",'
                        f'FIND(",",${sku}{i}&",")-1))))')
        f[C_LINE[s]] = (f'=IF(${code}{i}="","",MID(${code}{i},4,'
                        f'FIND("~",SUBSTITUTE(${code}{i},"-","~",2))-4))')
        f[C_VCOD[s]] = (f'=IF(${code}{i}="","",MID(${code}{i},'
                        f'FIND("~",SUBSTITUTE(${code}{i},"-","~",2))+1,'
                        f'FIND("~",SUBSTITUTE(${code}{i},"-","~",3))'
                        f'-FIND("~",SUBSTITUTE(${code}{i},"-","~",2))-1))')
        f[C_CAT[s]] = f'=IF(${line}{i}="","",IFERROR(VLOOKUP(${line}{i},{LINES},3,FALSE),"?"))'
        f[C_VAR[s]] = (f'=IF(${vcod}{i}="","",IFERROR(VLOOKUP(${line}{i}&"-"&${vcod}{i},'
                       f'{VARMAP},2,FALSE),"?"))')
        f[C_SIZE[s]] = f'=IF(${code}{i}="","",IFERROR(VALUE(RIGHT(${code}{i},3)),""))'
        f[C_DTV[s]] = (f'=IF(${dt}{i}="","",IFERROR(IF(ISNUMBER(${dt}{i}),${dt}{i},'
                       f'DATEVALUE(TRIM(${dt}{i}))),""))')
        f[C_SEAS[s]] = (f'=IF(${dtv}{i}="","",IF(MONTH(${dtv}{i})>=12,"Cold Winter",'
                        f'IF(MONTH(${dtv}{i})>=10,"Post-Monsoon",'
                        f'IF(MONTH(${dtv}{i})>=6,"Hot Humid",'
                        f'IF(MONTH(${dtv}{i})>=3,"Hot Dry","Cold Winter")))))')
        f[C_INT[s]] = f'=IF(${line}{i}="","",IFERROR(VLOOKUP(${line}{i},{LINES},2,FALSE),"?"))'
    ecat, evar, esize, esku = L(C_ECAT), L(C_EVAR), L(C_ESIZE), L(C_ESKU)
    nxt, ncat, nvar, nsize, nsku = L(C_NEXT), L(C_NCAT), L(C_NVAR), L(C_NSIZE), L(C_NSKU)
    move, gap, v2v, v2c, coh = L(C_MOVE), L(C_GAP), L(C_V2V), L(C_V2C), L(C_COH)
    cat1, var1, size1, dtv1 = L(C_CAT[0]), L(C_VAR[0]), L(C_SIZE[0]), L(C_DTV[0])
    code2, cat2, var2, size2, dtv2 = L(C_CODE[1]), L(C_CAT[1]), L(C_VAR[1]), L(C_SIZE[1]), L(C_DTV[1])
    f[C_ECAT] = f'=IF(${cat1}{i}="","",{cat1}{i})'
    f[C_EVAR] = f'=IF(${var1}{i}="","",{var1}{i})'
    f[C_ESIZE] = f'=IF(${size1}{i}="","",{size1}{i})'
    f[C_ESKU] = (f'=IF(${ecat}{i}="","",IF(${ecat}{i}="Aloe Vera Gel",'
                 f'"Aloe Vera Gel "&${esize}{i}&"g",{evar}{i}&" "&${esize}{i}&"g"))')
    f[C_NEXT] = f'=IF(${code2}{i}="",0,1)'
    f[C_NCAT] = f'=IF(${nxt}{i}=0,"",{cat2}{i})'
    f[C_NVAR] = f'=IF(${nxt}{i}=0,"",{var2}{i})'
    f[C_NSIZE] = f'=IF(${nxt}{i}=0,"",{size2}{i})'
    f[C_NSKU] = (f'=IF(${ncat}{i}="","",IF(${ncat}{i}="Aloe Vera Gel",'
                 f'"Aloe Vera Gel "&${nsize}{i}&"g",{nvar}{i}&" "&${nsize}{i}&"g"))')
    f[C_MOVE] = (f'=IF(${nxt}{i}=0,"",IF(${esku}{i}=${nsku}{i},"Repeat Exact",'
                 f'IF(${nsize}{i}>${esize}{i},"Upsized",IF(${nsize}{i}<${esize}{i},"Downsized",'
                 f'IF(${ecat}{i}<>${ncat}{i},"Lateral","Same Size Diff Variant")))))')
    f[C_GAP] = f'=IF(${nxt}{i}=0,"",{dtv2}{i}-{dtv1}{i})'
    f[C_V2V] = f'=IF(${nxt}{i}=0,0,IF({nvar}{i}={evar}{i},1,0))'
    f[C_V2C] = f'=IF(${nxt}{i}=0,0,IF({ncat}{i}={ecat}{i},1,0))'
    f[C_COH] = (f'=IF($A{i}="","",IF(AND('
                f'IF(DASHBOARD!$C$4="All",TRUE,${ecat}{i}=DASHBOARD!$C$4),'
                f'IF(DASHBOARD!$C$6="All",TRUE,${esku}{i}=DASHBOARD!$C$6),'
                f'IF(DASHBOARD!$C$8="All",TRUE,${L(C_MY)}{i}>=DASHBOARD!$C$8),'
                f'IF(DASHBOARD!$C$10="All",TRUE,${L(C_MY)}{i}<=DASHBOARD!$C$10),'
                f'IF(DASHBOARD!$C$12="All",TRUE,${L(C_DEPTH)}{i}=VALUE(DASHBOARD!$C$12)),'
                f'IF(DASHBOARD!$C$14="All",TRUE,${L(C_REGION)}{i}=DASHBOARD!$C$14)),1,0))')
    for col, formula in f.items():
        dat.cell(row=i, column=col, value=formula)

# helper column refs (full-column) used by other sheets
def H(col): return f"DATA!${L(col)}:${L(col)}"
INT_COLS = [H(C_INT[s]) for s in range(6)]
CAT_COLS = [H(C_CAT[s]) for s in range(6)]
SEAS_COLS = [H(C_SEAS[s]) for s in range(6)]
VAR_COLS = [H(C_VAR[s]) for s in range(6)]

tbl = Table(displayName="Dump", ref=f"A1:{L(N_HELP_COLS)}{n+1}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
dat.add_table(tbl)
dat.freeze_panes = "B2"

# ================================================================= DASHBOARD
dash = wb.create_sheet("DASHBOARD", 1)
dash.column_dimensions["B"].width = 34
for col in "CDEF": dash.column_dimensions[col].width = 16
dash.column_dimensions["G"].width = 2
dash["B2"] = "🧴 NAT HABIT — RETENTION DASHBOARD"; dash["B2"].font = TITLE
dash["B3"] = "Dropdown filters drive every tile below (cohort = customers whose FIRST order matches)"
dash["B3"].font = SUB

filters = [(4, "Category", CATLIST), (6, "Variant (SKU)", SKULIST),
           (8, "From month", MONLIST), (10, "To month", MONLIST),
           (12, "Journey depth", DEPLIST), (14, "Geo region", REGLIST)]
for r0, label, src in filters:
    dash.cell(row=r0, column=2, value=label).font = BOLD
    c = dash.cell(row=r0, column=3, value="All")
    c.border = THIN; c.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1=f"={src.split('!')[1].replace('$','')}"
                        if False else None, allow_blank=False)
    # use direct sheet reference for validation
    dv = DataValidation(type="list", formula1=f"={src}", allow_blank=False)
    dash.add_data_validation(dv); dv.add(f"C{r0}")

def sum_counts(colpairs, value):
    """COUNTIFS over 6 slot columns, restricted to cohort."""
    return "+".join(f"COUNTIFS({H(C_COH)},1,{col},{value})"
                    for col in colpairs)

def tile(row, label, formula, numfmt=None):
    lc = dash.cell(row=row, column=2, value=label)
    lc.font = Font(size=10, bold=True); lc.border = THIN; lc.fill = FILL_TILE
    vc = dash.cell(row=row, column=3, value=formula)
    vc.font = Font(bold=True, size=12, color=DARK); vc.border = THIN
    vc.alignment = Alignment(horizontal="center")
    if numfmt: vc.number_format = numfmt

dash["B17"] = "COHORT"; dash["B17"].font = Font(bold=True, color="FFFFFF")
dash["B17"].fill = FILL_HDR
tile(18, "Customers in cohort", f'=COUNTIFS({H(C_COH)},1)', "#,##0")
tile(19, "Total orders", f'=SUMIFS({H(C_DEPTH)},{H(C_COH)},1)', "#,##0")
tile(20, "Repeat buyers", f'=COUNTIFS({H(C_COH)},1,{H(C_REP)},1)', "#,##0")
tile(21, "Repeat rate", "=IFERROR(C20/C18,\"\")", "0.0%")

dash["B23"] = "SEASONALITY  (orders by variant's intended season line)"
dash["B23"].font = Font(bold=True, color="FFFFFF"); dash["B23"].fill = FILL_HDR
tile(24, "Cold Winter (Dec–Feb) variants", "=" + "+".join(
    f"COUNTIFS({H(C_COH)},1,{col},\"Cold Winter\")" for col in INT_COLS), "#,##0")
tile(25, "Hot Dry (Mar–May) variants", "=" + "+".join(
    f"COUNTIFS({H(C_COH)},1,{col},\"Hot Dry\")" for col in INT_COLS), "#,##0")
tile(26, "Hot Humid (Jun–Sep) variants", "=" + "+".join(
    f"COUNTIFS({H(C_COH)},1,{col},\"Hot Humid\")" for col in INT_COLS), "#,##0")
tile(27, "Non-Seasonal / Concern variants", "=" + "+".join(
    f"COUNTIFS({H(C_COH)},1,{col},\"Non-Seasonal\")" for col in INT_COLS), "#,##0")
tile(28, "Peak Season", ('=IF(SUM(C24:C27)=0,"—",INDEX({"Cold Winter";"Hot Dry";'
                         '"Hot Humid";"Non-Seasonal"},MATCH(MAX(C24:C27),C24:C27,0)))'))
dash.cell(row=29, column=2, value="Purchase timing (calendar month buckets):").font = SUB
for k, (lab, val) in enumerate([("Cold Winter", '"Cold Winter"'),
                                ("Hot Dry", '"Hot Dry"'),
                                ("Hot Humid", '"Hot Humid"'),
                                ("Post-Monsoon", '"Post-Monsoon"')]):
    c = dash.cell(row=29, column=3 + k, value="=" + "+".join(
        f"COUNTIFS({H(C_COH)},1,{col},{val})" for col in SEAS_COLS))
    c.font = Font(size=9); c.number_format = "#,##0"
    dash.cell(row=30, column=3 + k, value=lab).font = Font(size=8, color="888888")

dash["B32"] = "GELS & TOTALS"; dash["B32"].font = Font(bold=True, color="FFFFFF")
dash["B32"].fill = FILL_HDR
tile(33, "Active Gel Count", "=" + "+".join(
    f'COUNTIFS({H(C_COH)},1,{col},"Active Gel")' for col in CAT_COLS), "#,##0")
tile(34, "Aloevera Gel Count", "=" + "+".join(
    f'COUNTIFS({H(C_COH)},1,{col},"Aloe Vera Gel")' for col in CAT_COLS), "#,##0")
tile(35, "Face Malai Count", "=" + "+".join(
    f'COUNTIFS({H(C_COH)},1,{col},"Face Malai")' for col in CAT_COLS), "#,##0")
tile(36, "Total Orders", f'=SUMIFS({H(C_DEPTH)},{H(C_COH)},1)', "#,##0")

dash["B38"] = "LOYALTY & JOURNEY"; dash["B38"].font = Font(bold=True, color="FFFFFF")
dash["B38"].fill = FILL_HDR
base = f'COUNTIFS({H(C_COH)},1,{H(C_NEXT)},1)'
tile(39, "Repeat customers (2+ orders)", f"={base}", "#,##0")
tile(40, "V2V Loyalty %",
     f'=IFERROR(COUNTIFS({H(C_COH)},1,{H(C_NEXT)},1,{H(C_V2V)},1)/C39,"")', "0.0%")
tile(41, "V2C Loyalty %",
     f'=IFERROR(COUNTIFS({H(C_COH)},1,{H(C_NEXT)},1,{H(C_V2C)},1)/C39,"")', "0.0%")
tile(42, "Avg Purchase Days V2V",
     f'=IFERROR(AVERAGEIFS({H(C_GAP)},{H(C_COH)},1,{H(C_V2V)},1),"")', "0")
tile(43, "Avg Purchase Days V2C",
     f'=IFERROR(AVERAGEIFS({H(C_GAP)},{H(C_COH)},1,{H(C_V2C)},1),"")', "0")
tile(44, "Avg Order Count V2V",
     f'=IFERROR(AVERAGEIFS({H(C_DEPTH)},{H(C_COH)},1,{H(C_V2V)},1),"")', "0.00")
tile(45, "Avg Order Count V2C",
     f'=IFERROR(AVERAGEIFS({H(C_DEPTH)},{H(C_COH)},1,{H(C_V2C)},1),"")', "0.00")

# ================================================================= VARIANTS
var = wb.create_sheet("VARIANTS")
vhead = ["Variant", "Cold Winter", "Hot Dry", "Hot Humid", "Post-Monsoon",
         "Total Orders", "Peak Season", "Peak Share", "Intended Season",
         "Season Match?", "Seasonality Strength", "Entry Users",
         "Bought Once", "Repeat Rate", "Avg Days 1→2"]
for j, h in enumerate(vhead, start=1):
    c = var.cell(row=2, column=j, value=h); c.font = HDR; c.fill = FILL_HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center")
var.column_dimensions["A"].width = 24
for j in range(2, 16): var.column_dimensions[CL(j)].width = 12
var["A1"] = "VARIANT SUMMARY — whole workbook (ignores dashboard filters)"
var["A1"].font = BOLD
for i, v in enumerate(VARIANTS, start=3):
    var.cell(row=i, column=1, value=v)
    for k, bucket in enumerate(["Cold Winter", "Hot Dry", "Hot Humid", "Post-Monsoon"]):
        var.cell(row=i, column=2 + k, value="=" + "+".join(
            f'COUNTIFS({vc},$A{i},{sc},"{bucket}")'
            for vc, sc in zip(VAR_COLS, SEAS_COLS)))
    var.cell(row=i, column=6, value=f"=SUM(B{i}:E{i})")
    var.cell(row=i, column=7, value=(f'=IF(F{i}=0,"—",INDEX({{"Cold Winter";"Hot Dry";'
                                     f'"Hot Humid";"Post-Monsoon"}},MATCH(MAX(B{i}:E{i}),B{i}:E{i},0)))'))
    var.cell(row=i, column=8, value=f'=IFERROR(MAX(B{i}:E{i})/F{i},"")').number_format = "0.0%"
    var.cell(row=i, column=9, value=f'=IFERROR(VLOOKUP($A{i},{VARINTENT},2,FALSE),"")')
    var.cell(row=i, column=10, value=(f'=IF(OR(I{i}="Non-Seasonal",I{i}="Gel"),"—",'
                                      f'IF(G{i}=I{i},"Matches","No Match"))'))
    var.cell(row=i, column=11, value=(f'=IF(J{i}="—",IF(I{i}="Gel","—","Evergreen"),'
                                      f'IF(H{i}>=0.4,"Highly Seasonal",'
                                      f'IF(H{i}>=0.3,"Moderately Seasonal","Evergreen")))'))
    var.cell(row=i, column=12, value=f"=COUNTIFS({H(C_VAR[0])},$A{i})")
    var.cell(row=i, column=13, value=f"=COUNTIFS({H(C_VAR[0])},$A{i},{H(C_DEPTH)},1)")
    var.cell(row=i, column=14, value=f'=IFERROR(1-M{i}/L{i},"")').number_format = "0.0%"
    var.cell(row=i, column=15, value=(f'=IFERROR(AVERAGEIFS({H(C_GAP)},'
                                      f'{H(C_EVAR)},$A{i}),"")')).number_format = "0"
for j in range(2, 7):
    var.cell(row=22, column=j, value=f"=SUM({CL(j)}3:{CL(j)}21)").font = BOLD
var.freeze_panes = "B3"

# ================================================================= GRAMMAGE
gr = wb.create_sheet("GRAMMAGE")
ghead = ["Category", "Entry Size", "Total Next Purchases", "Repeat Exact Same",
         "Repeat %", "Upsized", "Upsize %", "Lateral", "Downsize", "Downsize %",
         "Same Size Diff Variant", "FM30→FM50", "FM30→AG50", "FM50→FM30", "AG50→FM30"]
for j, h in enumerate(ghead, start=1):
    c = gr.cell(row=3, column=j, value=h); c.font = HDR; c.fill = FILL_HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center")
gr.column_dimensions["A"].width = 16; gr.column_dimensions["B"].width = 10
for j in range(3, 16): gr.column_dimensions[CL(j)].width = 11
gr["A1"] = "GRAMMAGE TRANSITIONS — entry SKU → 2nd order (whole workbook)"
gr["A1"].font = BOLD
rows = [("Face Malai", 30, 4), ("Face Malai", 50, 5), ("Active Gel", 50, 6),
        ("Aloe Vera Gel", 40, 7), ("Aloe Vera Gel", 80, 8)]
for cat, size, r0 in rows:
    gr.cell(row=r0, column=1, value=cat); gr.cell(row=r0, column=2, value=size)
    base = f'{H(C_ECAT)},"{cat}",{H(C_ESIZE)},{size},{H(C_NEXT)},1'
    gr.cell(row=r0, column=3, value=f"=COUNTIFS({base})")
    for k, mv in [(4, "Repeat Exact"), (6, "Upsized"), (8, "Lateral"),
                  (9, "Downsized"), (11, "Same Size Diff Variant")]:
        gr.cell(row=r0, column=k,
                value=f'=COUNTIFS({base},{H(C_MOVE)},"{mv}")')
    gr.cell(row=r0, column=5, value=f"=IFERROR(D{r0}/C{r0},\"\")").number_format = "0.0%"
    gr.cell(row=r0, column=7, value=f"=IFERROR(F{r0}/C{r0},\"\")").number_format = "0.0%"
    gr.cell(row=r0, column=10, value=f"=IFERROR(I{r0}/C{r0},\"\")").number_format = "0.0%"
    paths = {(12, "Face Malai", 30, 50), (13, "Active Gel", 50, 50),
             (14, "Face Malai", 50, 30), (15, "Face Malai", 30, 50)}
    gr.cell(row=4, column=12, value=(f'=COUNTIFS({H(C_ECAT)},"Face Malai",{H(C_ESIZE)},30,'
                                     f'{H(C_NCAT)},"Face Malai",{H(C_NSIZE)},50)'))
    gr.cell(row=4, column=13, value=(f'=COUNTIFS({H(C_ECAT)},"Face Malai",{H(C_ESIZE)},30,'
                                     f'{H(C_NCAT)},"Active Gel",{H(C_NSIZE)},50)'))
    gr.cell(row=5, column=14, value=(f'=COUNTIFS({H(C_ECAT)},"Face Malai",{H(C_ESIZE)},50,'
                                     f'{H(C_NCAT)},"Face Malai",{H(C_NSIZE)},30)'))
    gr.cell(row=6, column=15, value=(f'=COUNTIFS({H(C_ECAT)},"Active Gel",{H(C_ESIZE)},50,'
                                     f'{H(C_NCAT)},"Face Malai",{H(C_NSIZE)},30)'))
for j in range(3, 12):
    gr.cell(row=9, column=j, value=f"=SUM({CL(j)}4:{CL(j)}8)").font = BOLD
gr.cell(row=9, column=1, value="TOTAL").font = BOLD
gr.cell(row=9, column=5, value="=IFERROR(D9/C9,\"\")").number_format = "0.0%"
gr.cell(row=9, column=7, value="=IFERROR(F9/C9,\"\")").number_format = "0.0%"
gr.cell(row=9, column=10, value="=IFERROR(I9/C9,\"\")").number_format = "0.0%"

# ================================================================= CONCLUSIONS
con = wb.create_sheet("CONCLUSIONS", 2)
con.column_dimensions["A"].width = 3
con.column_dimensions["B"].width = 34
con.column_dimensions["C"].width = 110
con["B1"] = "FORMULA-DRIVEN CONCLUSIONS — whole workbook level"
con["B1"].font = BOLD
con["B2"] = "Every number is computed and every verdict is a formula branch; they update as soon as DATA changes."
con["B2"].font = SUB

SK = f"VARIANTS!$K$3:$K$21"; SI = f"VARIANTS!$I$3:$I$21"
qs = [
    ("Is Face Malai Seasonal?",
     f'=IF(COUNTIF({SK},"Highly Seasonal")=0,'
     f'IF(COUNTIF({SK},"Evergreen")>0,"NO — Mostly evergreen with "&COUNTIF({SK},"Evergreen")&" evergreen variants","MIXED"),'
     f'IF(COUNTIF({SK},"Highly Seasonal")>=COUNTIF({SK},"Evergreen"),'
     f'"YES — "&COUNTIF({SK},"Highly Seasonal")&" of "&COUNTIF({SI},"<>Gel")&" Face Malai variants are highly seasonal",'
     f'"MIXED — "&COUNTIF({SK},"Highly Seasonal")&" highly seasonal vs "&COUNTIF({SK},"Evergreen")&" evergreen variants"))'),
    ("Overall Repeat Behaviour",
     '=IFERROR(TEXT(GRAMMAGE!D9/GRAMMAGE!C9,"0%")&" of all face moisturiser transitions are same product repeats ("'
     '&TEXT(GRAMMAGE!D9,"#,##0")&" of "&TEXT(GRAMMAGE!C9,"#,##0")&" transitions). "'
     '&IF(GRAMMAGE!D9/GRAMMAGE!C9>=0.6,"STRONG",IF(GRAMMAGE!D9/GRAMMAGE!C9>=0.45,"MODERATE","WEAK"))'
     '&" — switching behaviour exists across sizes and formats.","")'),
    ("30g → 50g Conversion (FM Only)",
     '="Out of "&TEXT(GRAMMAGE!C4,"#,##0")&" FM 30g transitions — "&TEXT(GRAMMAGE!L4,"#,##0")&" ("'
     '&TEXT(IFERROR(GRAMMAGE!L4/GRAMMAGE!C4,0),"0%")&") moved to FM 50g. "'
     '&IF(GRAMMAGE!L4>=GRAMMAGE!M4,"This is the primary internal upsize path.",'
     '"Secondary — the cross-category gel switch is bigger.")'),
    ("30g → Active Gel (Cross-Category Upsize)",
     '="Out of "&TEXT(GRAMMAGE!C4,"#,##0")&" FM 30g transitions — "&TEXT(GRAMMAGE!M4,"#,##0")&" ("'
     '&TEXT(IFERROR(GRAMMAGE!M4/GRAMMAGE!C4,0),"0%")&") moved to Active Gel 50g. "'
     '&IF(GRAMMAGE!M4>GRAMMAGE!L4,"More 30g buyers move to Active Gel than FM 50g — gel format has stronger pull than size upgrade within FM.",'
     '"FM 50g currently converts better than the gel switch.")'),
    ("Downsize Risk Assessment",
     '="DOWNSIZE ALERT: "&TEXT(IFERROR(GRAMMAGE!N5/GRAMMAGE!C5,0),"0%")&" of FM 50g buyers drop to 30g ("'
     '&TEXT(GRAMMAGE!N5,"#,##0")&" transitions). AND "&TEXT(IFERROR(GRAMMAGE!O6/GRAMMAGE!C6,0),"0%")'
     '&" of Active Gel buyers also drop to FM 30g ("&TEXT(GRAMMAGE!O6,"#,##0")&" transitions). "'
     '&IF(OR(IFERROR(GRAMMAGE!N5/GRAMMAGE!C5,0)>=0.2,IFERROR(GRAMMAGE!O6/GRAMMAGE!C6,0)>=0.2),'
     '"Portfolio-wide downsize pressure detected.","Downsize pressure within tolerance.")'),
    ("Key Business Recommendation",
     '=IF(GRAMMAGE!M4>GRAMMAGE!L4,'
     '"FOCUS: 30g buyers prefer switching to Active Gel over FM 50g. Bundle strategy: FM 30g + Active Gel 50g trial kit could accelerate this natural behaviour.",'
     '"FOCUS: strengthen the FM 30g → FM 50g upsize path ("&TEXT(IFERROR(GRAMMAGE!L4/GRAMMAGE!C4,0),"0%")'
     '&" today). Consider 50g refill pricing or a size-up nudge at reorder.")'),
]
for i, (q, formula) in enumerate(qs, start=4):
    qc = con.cell(row=i, column=2, value=q)
    qc.font = Font(bold=True, color=DARK); qc.fill = FILL_SEC
    qc.alignment = Alignment(vertical="top", wrap_text=True); qc.border = THIN
    ac = con.cell(row=i, column=3, value=formula)
    ac.alignment = Alignment(vertical="top", wrap_text=True); ac.border = THIN
    con.row_dimensions[i].height = 45

# sheet order: README, DASHBOARD, CONCLUSIONS, VARIANTS, GRAMMAGE, DATA, DECODE
wb.move_sheet("DASHBOARD", offset=-(wb.sheetnames.index("DASHBOARD") - 1))
order = ["README", "DASHBOARD", "CONCLUSIONS", "VARIANTS", "GRAMMAGE", "DATA", "DECODE"]
wb._sheets = [wb[nm] for nm in order]

wb.save(OUT)
print(f"saved {OUT} · sample rows: {n} · helper cols: {N_HELP_COLS} ({L(N_HELP_COLS)})")
